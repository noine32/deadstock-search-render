import pandas as pd
import chardet
import io
from datetime import datetime
from openpyxl.styles import Border, Side


class FileProcessor:

    @staticmethod
    def detect_encoding(file_bytes):
        result = chardet.detect(file_bytes)
        return result['encoding']

    @staticmethod
    def read_excel(file):
        try:
            df = pd.read_excel(file, engine='openpyxl')
            return df
        except Exception as e:
            raise Exception(f"Excelファイルの読み込みエラー: {str(e)}")

    @staticmethod
    def read_csv(file, file_type='default'):
        try:
            file_bytes = file.getvalue()
            encoding = FileProcessor.detect_encoding(file_bytes)

            if file_type == 'inventory':
                # 不良在庫データの場合、最初の7行をスキップ
                df = pd.read_csv(io.BytesIO(file_bytes),
                                 encoding=encoding,
                                 skiprows=7)

                # デバッグ用：元のデータ行数を記録
                print(f"読み込み直後の行数: {len(df)}")

                # 薬品名が空白の行を削除（より厳密なチェック）
                # NaN, None, 空文字、空白文字をすべて除外
                df['薬品名'] = df['薬品名'].astype(str).apply(lambda x: x.strip())
                df = df[df['薬品名'].apply(
                    lambda x: x not in ['', 'nan', 'None'])]

                # デバッグ用：薬品名フィルタリング後の行数を記録
                print(f"薬品名フィルタリング後の行数: {len(df)}")

                # 在庫量を数値に変換し、0以下の行を削除
                df['在庫量'] = pd.to_numeric(df['在庫量'], errors='coerce')
                df = df[df['在庫量'] > 0]

                # デバッグ用：在庫量フィルタリング後の行数を記録
                print(f"在庫量フィルタリング後の行数: {len(df)}")

                # 在庫量を整数に変換
                df['在庫量'] = df['在庫量'].astype(int)
            else:
                df = pd.read_csv(io.BytesIO(file_bytes), encoding=encoding)

            return df
        except Exception as e:
            raise Exception(f"CSVファイルの読み込みエラー: {str(e)}")

    @staticmethod
    def process_data(purchase_history_df, inventory_df, yj_code_df):
        try:
            print("データ処理開始")
            print(
                f"入力データの行数: 購入履歴={len(purchase_history_df)}, 在庫={len(inventory_df)}, YJコード={len(yj_code_df)}"
            )

            # データの前処理と検証
            # 空の薬品名を持つ行を削除
            inventory_df = inventory_df[inventory_df['薬品名'].notna() & (
                inventory_df['薬品名'].str.strip() != '')]
            print(f"薬品名フィルタリング後の在庫データ行数: {len(inventory_df)}")

            # 在庫量のバリデーション
            inventory_df['在庫量'] = pd.to_numeric(inventory_df['在庫量'],
                                                errors='coerce')
            inventory_df = inventory_df[inventory_df['在庫量'] > 0]
            print(f"在庫量バリデーション後の行数: {len(inventory_df)}")

            # 使用期限のフォーマットチェックと変換
            inventory_df['使用期限'] = pd.to_datetime(inventory_df['使用期限'],
                                                  errors='coerce')
            inventory_df = inventory_df[inventory_df['使用期限'].notna()]
            print(f"使用期限バリデーション後の行数: {len(inventory_df)}")

            # 数値データを文字列に変換し、NaN値を処理
            for df in [inventory_df, purchase_history_df, yj_code_df]:
                for col in df.columns:
                    df[col] = df[col].fillna('').astype(str)

            # 在庫金額CSVから薬品名とＹＪコードのマッピングを作成
            yj_mapping = dict(
                zip(yj_code_df['薬品名'],
                    zip(yj_code_df['ＹＪコード'], yj_code_df['単位'])))
            print(f"YJコードマッピング数: {len(yj_mapping)}")

            # 不良在庫データに対してＹＪコードと単位を設定
            inventory_df['ＹＪコード'] = inventory_df['薬品名'].map(
                lambda x: yj_mapping.get(x.strip(), (None, None))[0])
            inventory_df['単位'] = inventory_df['薬品名'].map(
                lambda x: yj_mapping.get(x.strip(), (None, None))[1])

            # マッピング結果の確認
            mapped_count = inventory_df['ＹＪコード'].notna().sum()
            print(f"YJコードマッピング成功数: {mapped_count}/{len(inventory_df)}")

            # ＹＪコードと厚労省CDで紐付け
            merged_df = pd.merge(inventory_df,
                                 purchase_history_df[[
                                     '厚労省CD', '法人名', '院所名', '品名・規格', '新薬品ｺｰﾄﾞ'
                                 ]],
                                 left_on='ＹＪコード',
                                 right_on='厚労省CD',
                                 how='left')
            print(f"マージ後のデータ行数: {len(merged_df)}")

            # 院所名別にデータを整理
            result_df = merged_df[[
                '品名・規格', '在庫量', '単位', '新薬品ｺｰﾄﾞ', '使用期限', 'ロット番号', '法人名', '院所名'
            ]].copy()

            # 空の値を空文字列に変換
            result_df = result_df.fillna('')

            # データの検証
            # 必須項目のチェック
            required_columns = ['品名・規格', '在庫量', '使用期限']
            for col in required_columns:
                missing = result_df[result_df[col] == ''].shape[0]
                print(f"{col}の欠損数: {missing}")

            # 院所名でソート
            result_df = result_df.sort_values(['法人名', '院所名'])
            print(f"最終データ行数: {len(result_df)}")

            return result_df

        except Exception as e:
            print(f"データ処理中にエラーが発生: {str(e)}")
            raise Exception(f"データ処理エラー: {str(e)}")

    @staticmethod
    def generate_excel(df):
        excel_buffer = io.BytesIO()

        # シート名として無効な文字を置換する関数
        def clean_sheet_name(name):
            if not isinstance(name, str) or not name.strip():
                return 'Unknown'
            # 特殊文字を置換
            invalid_chars = ['/', '\\', '?', '*', ':', '[', ']']
            cleaned_name = ''.join('_' if c in invalid_chars else c
                                   for c in name)
            # 最大31文字に制限（Excelの制限）
            return cleaned_name[:31].strip()

        # ExcelWriterを使用して、院所名ごとにシートを作成
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # ダミーのシートを最初に作成
            dummy_sheet_name = "dummy_sheet"  # ダミーシートの名前
            dummy_sheet = writer.book.create_sheet(dummy_sheet_name)
            sheet_created = False  # シートが作成されたかどうかを追跡するフラグ

            # 院所名ごとにシートを作成（空の値を除外）

            for name in df['院所名'].unique():
                if pd.notna(name) and str(name).strip():  # 空の値をスキップ
                    sheet_name = clean_sheet_name(str(name))
                    # 該当する院所のデータを抽出
                    sheet_df = df[df['院所名'] == name].copy()

                    # シートを定義（sheet_dfが空でも定義されるように）
                    worksheet = writer.book.create_sheet(sheet_name)

                    if not sheet_df.empty:
                        # 法人名と院所名を取得（ヘッダー用）
                        houjin_name = sheet_df['法人名'].iloc[0]
                        insho_name = sheet_df['院所名'].iloc[0]

                        # 以下の部分を修正
                        houjin_name = str(houjin_name).strip() if pd.notna(
                            houjin_name) else ''
                        insho_name = str(insho_name).strip() if pd.notna(
                            insho_name) else ''

                        print(f"houjin_name: {houjin_name}")
                        print(f"insho_name: {insho_name}")

                        if houjin_name and insho_name:  # 両方とも有効な場合にのみフォーマットを実行
                            header_text = '{} {} 御中'.format(
                                houjin_name, insho_name)
                        else:
                            header_text = ' 御中'  # 例：片方が無効な場合はデフォルトテキスト

                        if not sheet_df.empty:
                            # 表示用のカラムから法人名と院所名を除外
                            display_df = sheet_df.drop(['法人名', '院所名'], axis=1)

                            if not display_df.empty:  # display_df が空でない場合にのみ、シートを作成する
                                # 「引取り可能数」列を追加
                                display_df.insert(
                                    display_df.columns.get_loc('ロット番号') + 1,
                                    '引取り可能数', '')

                                # ヘッダー情報を作成
                                header_data = [
                                    ['不良在庫引き取り依頼'], [''], [header_text], [''],
                                    [
                                        '下記の不良在庫につきまして、引き取りのご検討を賜れますと幸いです。どうぞよろしくお願いいたします。'
                                    ], ['']
                                ]

                                # ヘッダー情報を worksheet.cell を使って直接書き込む
                                for row_num, row in enumerate(header_data):
                                    for col_num, cell in enumerate(row):
                                        worksheet.cell(row=row_num + 1,
                                                       column=col_num + 1,
                                                       value=cell)

                                display_df.to_excel(writer,
                                                    sheet_name=sheet_name,
                                                    startrow=6,
                                                    index=False)
                                sheet_created = True  # シート作成フラグをTrueにする
                            else:
                                print(
                                    f"display_dfが空なのでシート{sheet_name}は作成をスキップします"
                                )
                                # 空のデータフレームの場合、何もせず次のループへ
                                continue
                        else:
                            print(f"sheet_dfが空なのでシート{sheet_name}は作成をスキップします")
                            continue

                    # シートを取得してフォーマットを設定
                    # worksheet = writer.sheets[sheet_name] # この行は削除します

                    from openpyxl.styles import Border, Side
                    # 列幅の設定
                    worksheet.column_dimensions[
                        'A'].width = 35  # 255ピクセルは約35文字幅

                    # データの開始行（ヘッダーの後）
                    data_start_row = 7

                    # B～G列の幅を125ピクセル（約17文字幅）に設定
                    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
                        worksheet.column_dimensions[col].width = 17

                    # 罫線スタイルの定義
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

                    # データ部分に罫線を追加
                    for row in worksheet.iter_rows(min_row=data_start_row):
                        for cell in row:
                            cell.border = thin_border

                    # 行の高さを設定（30ピクセル）
                    for row in range(1, worksheet.max_row + 1):
                        worksheet.row_dimensions[row].height = 30

                    # すべてのセルのフォントサイズを14に設定
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.font is None:
                                cell.font = cell.font.copy()
                            cell.font = cell.font.copy(size=14)

                    # 特定のセルのフォント設定を上書き
                    cell_a1 = worksheet['A1']
                    cell_a1.font = cell_a1.font.copy(size=16)

                    cell_a3 = worksheet['A3']
                    cell_a3.font = cell_a3.font.copy(size=14, bold=True)

                    # 印刷設定
                    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
                    worksheet.print_title_rows = '1:7'  # 1-7行目をタイトル行として設定
                    worksheet.page_setup.fitToPage = True
                    worksheet.page_setup.fitToHeight = 0  # 高さは自動
                    worksheet.page_setup.fitToWidth = 1  # 幅は1ページに収める

            if not sheet_created:  # ループ内でシートが一つも作成されなかった場合はダミーシートを削除
                del writer.book[dummy_sheet_name]
        excel_buffer.seek(0)
        return excel_buffer
